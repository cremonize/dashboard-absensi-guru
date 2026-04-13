# ============================================================
#  DASHBOARD ABSENSI GURU — SMP Plus YPP Darussurur
#  Dibuat dengan Streamlit + Pandas + Plotly
#  Bisa dijalankan di: Lokal | Google Colab | Streamlit Cloud
# ============================================================

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import datetime

# ── Konfigurasi halaman ──────────────────────────────────────
st.set_page_config(
    page_title="Dashboard Absensi Guru",
    page_icon="🏫",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS Kustom ───────────────────────────────────────────────
st.markdown("""
<style>
    /* Kartu metrik */
    [data-testid="metric-container"] {
        background: linear-gradient(135deg, #1e3a5f 0%, #2d6a9f 100%);
        border-radius: 12px;
        padding: 16px;
        border-left: 5px solid #4fc3f7;
        box-shadow: 0 4px 15px rgba(0,0,0,0.2);
    }
    [data-testid="metric-container"] label { color: #b0d4f1 !important; font-size: 13px !important; }
    [data-testid="metric-container"] [data-testid="stMetricValue"] { color: #ffffff !important; font-size: 28px !important; font-weight: 700; }
    [data-testid="metric-container"] [data-testid="stMetricDelta"] { color: #81d4fa !important; }

    /* Header utama */
    .main-header {
        background: linear-gradient(135deg, #0d2137 0%, #1a4a7a 100%);
        padding: 24px 32px;
        border-radius: 16px;
        margin-bottom: 24px;
        border-bottom: 3px solid #4fc3f7;
    }
    .main-header h1 { color: #ffffff; font-size: 26px; margin: 0; font-weight: 700; }
    .main-header p  { color: #90caf9; font-size: 14px; margin: 6px 0 0 0; }

    /* Lencana status */
    .badge-hijau  { background:#1b5e20; color:#a5d6a7; padding:3px 10px; border-radius:20px; font-size:12px; font-weight:600; }
    .badge-kuning { background:#f57f17; color:#fff9c4; padding:3px 10px; border-radius:20px; font-size:12px; font-weight:600; }
    .badge-merah  { background:#b71c1c; color:#ffcdd2; padding:3px 10px; border-radius:20px; font-size:12px; font-weight:600; }

    /* Sidebar */
    section[data-testid="stSidebar"] { background: #0d1b2a !important; }
    section[data-testid="stSidebar"] * { color: #cfe2f3 !important; }
</style>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
#  FUNGSI PEMBACA DATA
# ════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def baca_excel(file_obj):
    """
    Membaca semua sheet dari file Excel absensi guru.
    Mengembalikan dict berisi DataFrame masing-masing sheet.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        data = {}

        # ── CONFIG ──────────────────────────────────────────
        ws = wb["CONFIG"]
        config = {}
        for row in ws.iter_rows(values_only=True):
            if row[1] and row[2]:
                config[str(row[1]).strip()] = row[2]
        data["config"] = config

        # ── DATA GURU ────────────────────────────────────────
        ws = wb["DATA_GURU"]
        rows = list(ws.iter_rows(values_only=True))
        guru_list = []
        for row in rows[2:]:  # skip judul & header
            if row[1] and row[2]:
                guru_list.append({"No": row[1], "Nama": row[2], "Mapel": row[3]})
        data["guru"] = pd.DataFrame(guru_list)

        # ── DATA JAM PELAJARAN ───────────────────────────────
        ws = wb["DATA_JAM_PELAJARAN"]
        rows = list(ws.iter_rows(values_only=True))
        jam_list = []
        for row in rows[3:]:  # skip 3 baris header
            if row[1] and row[2]:
                jam_list.append({
                    "No": row[1], "Nama": row[2], "Mapel": row[3],
                    "Senin": row[4] or 0, "Selasa": row[5] or 0,
                    "Rabu": row[6] or 0, "Kamis": row[7] or 0,
                    "Jumat": row[8] or 0, "Sabtu": row[9] or 0,
                    "Total_Minggu": row[10] or 0,
                })
        data["jam"] = pd.DataFrame(jam_list)

        # ── ABSENSI (data harian mentah) ─────────────────────
        ws = wb["ABSENSI"]
        rows = list(ws.iter_rows(values_only=True))

        # Baris 2 = tanggal, Baris 3 = nama hari, Baris 4+ = data guru
        tanggal_row = rows[1]   # index 1 → Row 2
        hari_row    = rows[2]   # index 2 → Row 3

        # Kumpulkan tanggal yang valid (kolom D dan seterusnya = index 3+)
        tanggal_cols = {}
        for col_idx in range(3, len(tanggal_row)):
            tgl = tanggal_row[col_idx]
            hari = hari_row[col_idx] if col_idx < len(hari_row) else None
            if tgl and hari and hari not in ("MINGGU", None):
                tanggal_cols[col_idx] = {"tanggal": tgl, "hari": hari}

        # Data absensi per guru per hari
        absensi_records = []
        for row in rows[3:]:  # rows index 3+ = data guru
            if not row[1]:
                continue
            nama_mapel = str(row[2]) if row[2] else ""
            nama = nama_mapel.split(" [")[0].strip() if " [" in nama_mapel else nama_mapel
            for col_idx, info in tanggal_cols.items():
                status = row[col_idx] if col_idx < len(row) else None
                absensi_records.append({
                    "Nama": nama,
                    "Tanggal": info["tanggal"],
                    "Hari": info["hari"],
                    "Status": status or "Tidak Hadir",
                })
        data["absensi"] = pd.DataFrame(absensi_records)
        if not data["absensi"].empty:
            data["absensi"]["Tanggal"] = pd.to_datetime(data["absensi"]["Tanggal"])

        # ── REKAP BULANAN ────────────────────────────────────
        ws = wb["REKAP_BULANAN"]
        rows = list(ws.iter_rows(values_only=True))
        rekap_list = []
        for row in rows[6:]:  # mulai row 7 (index 6)
            if row[1] and row[2] and str(row[2]) != "TOTAL":
                rekap_list.append({
                    "No":          row[1],
                    "Nama":        row[2],
                    "Mapel":       row[3],
                    "Target_Jam":  row[4] or 0,
                    "Jam_Hadir":   row[5] or 0,
                    "Jam_Sakit":   row[6] or 0,
                    "Jam_Izin":    row[7] or 0,
                    "Jam_TK":      row[8] or 0,
                    "Hari_Hadir":  row[10] or 0,
                    "Hari_Sakit":  row[11] or 0,
                    "Hari_Izin":   row[12] or 0,
                    "Hari_TK":     row[13] or 0,
                    "Pct_Hadir":   (row[14] or 0) * 100,
                    "Hari_Jadwal": row[15] or 0,
                })
        data["rekap"] = pd.DataFrame(rekap_list)

        return data, None

    except Exception as e:
        return None, str(e)


def hitung_rekap_dari_absensi(df_absensi, df_jam):
    """
    Menghitung ulang rekap dari data absensi mentah + jadwal jam.
    Digunakan ketika file CSV diupload (tanpa sheet REKAP).
    """
    HARI_JAM = {"Senin": "Senin", "Selasa": "Selasa", "Rabu": "Rabu",
                "Kamis": "Kamis", "Jumat": "Jumat", "Sabtu": "Sabtu"}
    jam_dict = df_jam.set_index("Nama").to_dict("index")
    rekap = []

    for nama, grp in df_absensi.groupby("Nama"):
        if nama not in jam_dict:
            continue
        jdwl = jam_dict[nama]
        target = jam_hadir = hari_hadir = hari_jadwal = 0
        for _, baris in grp.iterrows():
            hari = baris["Hari"]
            if hari not in HARI_JAM:
                continue
            jam_hari = jdwl.get(hari, 0)
            if jam_hari > 0:
                hari_jadwal += 1
                target += jam_hari
                if baris["Status"] == "Mengajar":
                    jam_hadir += jam_hari
                    hari_hadir += 1
        pct = (jam_hadir / target * 100) if target > 0 else 0
        rekap.append({
            "Nama": nama, "Target_Jam": target, "Jam_Hadir": jam_hadir,
            "Hari_Hadir": hari_hadir, "Hari_Jadwal": hari_jadwal, "Pct_Hadir": pct,
        })
    return pd.DataFrame(rekap)


def warna_pct(pct):
    if pct >= 90:
        return "badge-hijau", "✅ Baik"
    elif pct >= 75:
        return "badge-kuning", "⚠️ Cukup"
    else:
        return "badge-merah", "❌ Kurang"


# ════════════════════════════════════════════════════════════
#  SIDEBAR
# ════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## 🏫 Navigasi")
    halaman = st.radio(
        "Pilih Halaman",
        ["🏠 Beranda", "🏆 Ranking Guru", "👤 Detail Per Guru", "📅 Absensi Harian"],
        label_visibility="collapsed",
    )

    st.markdown("---")
    st.markdown("### 📂 Upload File")
    uploaded = st.file_uploader(
        "Upload file Excel (.xlsx)",
        type=["xlsx"],
        help="Upload file ABSENSI_GURU.xlsx untuk memperbarui data",
    )

    st.markdown("---")
    st.markdown("### ℹ️ Panduan Singkat")
    with st.expander("Cara pakai"):
        st.markdown("""
1. **Upload** file `.xlsx` absensi terbaru di atas
2. **Pilih halaman** untuk melihat laporan
3. **Export** tabel lewat tombol unduh
        """)
    st.markdown("---")
    st.caption("📌 Dashboard Absensi Guru v1.0")


# ════════════════════════════════════════════════════════════
#  MUAT DATA
# ════════════════════════════════════════════════════════════

# Coba baca dari upload; jika tidak ada, coba file lokal
_source = uploaded if uploaded else "ABSENSI_GURU_v4_fixed.xlsx"
data, err = baca_excel(_source)

if err or data is None:
    st.error(f"⚠️ Gagal membaca file. Pastikan format file benar.\n\nError: {err}")
    st.info("📤 Silakan upload file Excel absensi melalui sidebar kiri.")
    st.stop()

df_rekap   = data["rekap"]
df_jam     = data["jam"]
df_absensi = data["absensi"]
config     = data["config"]

# Info sekolah dari CONFIG
nama_sekolah = config.get("Nama Sekolah", "SMP Plus YPP Darussurur")
tgl_mulai    = config.get("Tanggal Mulai Periode", "")
tgl_akhir    = config.get("Tanggal Akhir Periode", "")
petugas      = config.get("Petugas / TTD", "")

if isinstance(tgl_mulai, datetime.datetime):
    tgl_mulai = tgl_mulai.strftime("%d %B %Y")
if isinstance(tgl_akhir, datetime.datetime):
    tgl_akhir = tgl_akhir.strftime("%d %B %Y")

periode_str = f"{tgl_mulai} s.d. {tgl_akhir}" if tgl_mulai else "—"

# ── Sort rekap untuk ranking ──────────────────────────────
df_rekap_sorted = df_rekap.sort_values("Pct_Hadir", ascending=False).reset_index(drop=True)
df_rekap_sorted["Rank"] = range(1, len(df_rekap_sorted) + 1)


# ════════════════════════════════════════════════════════════
#  HEADER BERSAMA
# ════════════════════════════════════════════════════════════

st.markdown(f"""
<div class="main-header">
  <h1>📊 Dashboard Kehadiran Guru</h1>
  <p>🏫 {nama_sekolah} &nbsp;|&nbsp; 📅 {periode_str} &nbsp;|&nbsp; 👤 {petugas}</p>
</div>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
#  HALAMAN 1 — BERANDA
# ════════════════════════════════════════════════════════════

if halaman == "🏠 Beranda":

    # ── Kartu Metrik ─────────────────────────────────────────
    total_guru    = len(df_rekap)
    total_target  = int(df_rekap["Target_Jam"].sum())
    total_hadir   = int(df_rekap["Jam_Hadir"].sum())
    rata_pct      = df_rekap["Pct_Hadir"].mean()
    guru_baik     = (df_rekap["Pct_Hadir"] >= 90).sum()
    guru_cukup    = ((df_rekap["Pct_Hadir"] >= 75) & (df_rekap["Pct_Hadir"] < 90)).sum()
    guru_kurang   = (df_rekap["Pct_Hadir"] < 75).sum()

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("👨‍🏫 Total Guru",   f"{total_guru} orang")
    c2.metric("🎯 Target Jam",     f"{total_target:,} jp",
              delta=f"Hadir: {total_hadir:,} jp")
    c3.metric("📈 Rata-rata Kehadiran", f"{rata_pct:.1f}%",
              delta="dari target jam")
    c4.metric("✅ Guru ≥ 90%", f"{guru_baik} guru",
              delta=f"⚠️ {guru_cukup}  ❌ {guru_kurang}")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Baris grafik ─────────────────────────────────────────
    col_pie, col_bar = st.columns([1, 2])

    with col_pie:
        st.subheader("📊 Distribusi Status Kehadiran")
        pie_data = pd.DataFrame({
            "Status": ["✅ Hadir (≥90%)", "⚠️ Cukup (75–89%)", "❌ Kurang (<75%)"],
            "Jumlah": [guru_baik, guru_cukup, guru_kurang],
        })
        fig_pie = px.pie(
            pie_data, names="Status", values="Jumlah",
            color="Status",
            color_discrete_map={
                "✅ Hadir (≥90%)": "#43a047",
                "⚠️ Cukup (75–89%)": "#fb8c00",
                "❌ Kurang (<75%)": "#e53935",
            },
            hole=0.45,
        )
        fig_pie.update_traces(textposition="inside", textinfo="percent+label",
                              textfont_size=12)
        fig_pie.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font_color="#e0e0e0", margin=dict(t=10, b=10),
            showlegend=False, height=320,
        )
        st.plotly_chart(fig_pie, use_container_width=True)

    with col_bar:
        st.subheader("📉 Jam Hadir vs Target Jam per Guru")
        df_bar = df_rekap_sorted[["Nama", "Target_Jam", "Jam_Hadir"]].copy()
        df_bar["Nama_Singkat"] = df_bar["Nama"].str.split(",").str[0]
        fig_bar = go.Figure()
        fig_bar.add_trace(go.Bar(
            name="Target Jam", x=df_bar["Nama_Singkat"], y=df_bar["Target_Jam"],
            marker_color="#1565c0", opacity=0.6,
        ))
        fig_bar.add_trace(go.Bar(
            name="Jam Hadir", x=df_bar["Nama_Singkat"], y=df_bar["Jam_Hadir"],
            marker_color="#42a5f5",
        ))
        fig_bar.update_layout(
            barmode="overlay", paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)", font_color="#e0e0e0",
            xaxis=dict(tickangle=-35, tickfont=dict(size=10)),
            yaxis_title="Jam Pelajaran", legend=dict(orientation="h"),
            margin=dict(t=10, b=10), height=320,
        )
        st.plotly_chart(fig_bar, use_container_width=True)

    # ── Tabel Ringkasan ───────────────────────────────────────
    st.subheader("📋 Ringkasan Semua Guru")
    df_tampil = df_rekap_sorted[["Rank", "Nama", "Mapel", "Target_Jam",
                                  "Jam_Hadir", "Hari_Hadir", "Pct_Hadir"]].copy()
    df_tampil.columns = ["#", "Nama Guru", "Mata Pelajaran",
                          "Target JP", "JP Hadir", "Hari Hadir", "% Hadir"]
    df_tampil["% Hadir"] = df_tampil["% Hadir"].map("{:.1f}%".format)

    st.dataframe(
        df_tampil, use_container_width=True, hide_index=True,
        column_config={
            "#": st.column_config.NumberColumn(width="small"),
            "% Hadir": st.column_config.TextColumn(width="medium"),
        }
    )

    # Tombol download
    buf = BytesIO()
    df_tampil.to_excel(buf, index=False)
    st.download_button(
        "⬇️ Unduh Rekap (Excel)", buf.getvalue(),
        file_name="rekap_kehadiran_guru.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ════════════════════════════════════════════════════════════
#  HALAMAN 2 — RANKING
# ════════════════════════════════════════════════════════════

elif halaman == "🏆 Ranking Guru":
    st.subheader("🏆 Ranking Guru Berdasarkan % Kehadiran Jam")

    # Filter interaktif
    col_f1, col_f2 = st.columns([1, 3])
    with col_f1:
        filter_status = st.selectbox(
            "Filter Status",
            ["Semua", "✅ ≥ 90% (Baik)", "⚠️ 75–89% (Cukup)", "❌ < 75% (Kurang)"],
        )
    with col_f2:
        mapel_list = ["Semua"] + sorted(df_rekap["Mapel"].unique().tolist())
        filter_mapel = st.selectbox("Filter Mata Pelajaran", mapel_list)

    df_rank = df_rekap_sorted.copy()
    if filter_status == "✅ ≥ 90% (Baik)":
        df_rank = df_rank[df_rank["Pct_Hadir"] >= 90]
    elif filter_status == "⚠️ 75–89% (Cukup)":
        df_rank = df_rank[(df_rank["Pct_Hadir"] >= 75) & (df_rank["Pct_Hadir"] < 90)]
    elif filter_status == "❌ < 75% (Kurang)":
        df_rank = df_rank[df_rank["Pct_Hadir"] < 75]
    if filter_mapel != "Semua":
        df_rank = df_rank[df_rank["Mapel"] == filter_mapel]

    # Grafik ranking horizontal
    df_rank_plot = df_rank.sort_values("Pct_Hadir")
    colors = ["#43a047" if p >= 90 else "#fb8c00" if p >= 75 else "#e53935"
              for p in df_rank_plot["Pct_Hadir"]]

    fig_rank = go.Figure(go.Bar(
        x=df_rank_plot["Pct_Hadir"],
        y=df_rank_plot["Nama"].str.split(",").str[0],
        orientation="h",
        marker_color=colors,
        text=df_rank_plot["Pct_Hadir"].map("{:.1f}%".format),
        textposition="outside",
    ))
    fig_rank.add_vline(x=90, line_dash="dash", line_color="#81d4fa",
                       annotation_text="Target 90%", annotation_font_color="#81d4fa")
    fig_rank.update_layout(
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font_color="#e0e0e0", xaxis=dict(range=[0, 110], title="% Kehadiran"),
        yaxis_title="", margin=dict(l=10, r=60, t=10, b=10),
        height=max(380, len(df_rank_plot) * 28),
    )
    st.plotly_chart(fig_rank, use_container_width=True)

    # Tabel ranking dengan badge
    st.markdown("#### 📄 Tabel Ranking Detail")
    for _, row in df_rank.iterrows():
        badge_cls, label = warna_pct(row["Pct_Hadir"])
        col1, col2, col3, col4, col5 = st.columns([0.5, 3, 1.5, 1.5, 1.5])
        col1.markdown(f"**#{int(row['Rank'])}**")
        col2.markdown(f"**{row['Nama']}**  \n*{row['Mapel']}*")
        col3.markdown(f"{int(row['Jam_Hadir'])} / {int(row['Target_Jam'])} JP")
        col4.markdown(f"{int(row['Hari_Hadir'])} / {int(row['Hari_Jadwal'])} Hari")
        col5.markdown(
            f"<span class='{badge_cls}'>{row['Pct_Hadir']:.1f}% {label}</span>",
            unsafe_allow_html=True,
        )
        st.divider()


# ════════════════════════════════════════════════════════════
#  HALAMAN 3 — DETAIL PER GURU
# ════════════════════════════════════════════════════════════

elif halaman == "👤 Detail Per Guru":
    st.subheader("👤 Detail Kehadiran Per Guru")

    nama_options = df_rekap_sorted["Nama"].tolist()
    nama_dipilih = st.selectbox("🔍 Pilih Nama Guru", nama_options)

    guru_row = df_rekap[df_rekap["Nama"] == nama_dipilih].iloc[0]

    # Kartu info guru
    badge_cls, label = warna_pct(guru_row["Pct_Hadir"])
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("🎯 Target JP",  int(guru_row["Target_Jam"]))
    m2.metric("✅ JP Hadir",   int(guru_row["Jam_Hadir"]))
    m3.metric("📅 Hari Hadir", int(guru_row["Hari_Hadir"]))
    m4.metric("📊 % Kehadiran", f"{guru_row['Pct_Hadir']:.1f}%")

    st.markdown(
        f"**Mata Pelajaran:** {guru_row['Mapel']} &nbsp;|&nbsp; "
        f"**Status:** <span class='{badge_cls}'>{label} ({guru_row['Pct_Hadir']:.1f}%)</span>",
        unsafe_allow_html=True,
    )

    # Jadwal jam per hari
    st.markdown("#### 📆 Jadwal Jam Mengajar per Hari")
    jam_row = df_jam[df_jam["Nama"] == nama_dipilih]
    if not jam_row.empty:
        j = jam_row.iloc[0]
        hari_cols = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu"]
        df_jadwal = pd.DataFrame({
            "Hari": hari_cols,
            "Jam Pelajaran": [int(j[h]) for h in hari_cols],
        })
        fig_jadwal = px.bar(
            df_jadwal, x="Hari", y="Jam Pelajaran",
            color="Jam Pelajaran", color_continuous_scale="Blues",
            text="Jam Pelajaran",
        )
        fig_jadwal.update_traces(textposition="outside")
        fig_jadwal.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font_color="#e0e0e0", showlegend=False, height=280,
            margin=dict(t=10, b=10), coloraxis_showscale=False,
        )
        st.plotly_chart(fig_jadwal, use_container_width=True)

    # Riwayat absensi
    st.markdown("#### 📋 Riwayat Absensi Harian")
    df_guru_abs = df_absensi[df_absensi["Nama"] == nama_dipilih].copy()
    if not df_guru_abs.empty:
        df_guru_abs = df_guru_abs.sort_values("Tanggal")
        df_guru_abs["Tanggal_Str"] = df_guru_abs["Tanggal"].dt.strftime("%d %b %Y")

        STATUS_ICON = {
            "Mengajar":            "🟢 Mengajar",
            "Sakit":               "🔴 Sakit",
            "Izin":                "🟡 Izin",
            "Tanpa Keterangan":    "⚫ Tanpa Ket.",
            "Dinas Keluar Sekolah":"🔵 Dinas",
            "Libur":               "⬜ Libur",
            "Libur Tanggal Merah": "🔶 Libur Merah",
        }
        df_guru_abs["Status_Label"] = df_guru_abs["Status"].map(
            lambda s: STATUS_ICON.get(s, f"⚪ {s}"))

        st.dataframe(
            df_guru_abs[["Tanggal_Str", "Hari", "Status_Label"]].rename(columns={
                "Tanggal_Str": "Tanggal", "Status_Label": "Status",
            }),
            use_container_width=True, hide_index=True,
        )
    else:
        st.info("Belum ada data absensi untuk guru ini.")

    # Diagram pie status guru ini
    if not df_guru_abs.empty:
        st.markdown("#### 🍩 Komposisi Status")
        count_status = df_guru_abs["Status"].value_counts().reset_index()
        count_status.columns = ["Status", "Jumlah"]
        fig_gpie = px.pie(
            count_status, names="Status", values="Jumlah", hole=0.4,
            color_discrete_sequence=px.colors.qualitative.Set2,
        )
        fig_gpie.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", font_color="#e0e0e0",
            margin=dict(t=10, b=10), height=280,
        )
        st.plotly_chart(fig_gpie, use_container_width=True)


# ════════════════════════════════════════════════════════════
#  HALAMAN 4 — ABSENSI HARIAN
# ════════════════════════════════════════════════════════════

elif halaman == "📅 Absensi Harian":
    st.subheader("📅 Data Absensi Harian")

    if df_absensi.empty:
        st.warning("Data absensi harian tidak tersedia.")
    else:
        # Filter tanggal
        tgl_min = df_absensi["Tanggal"].min().date()
        tgl_max = df_absensi["Tanggal"].max().date()

        col_d1, col_d2 = st.columns(2)
        with col_d1:
            tgl_awal = st.date_input("📅 Dari Tanggal", tgl_min,
                                     min_value=tgl_min, max_value=tgl_max)
        with col_d2:
            tgl_akhir_d = st.date_input("📅 Sampai Tanggal", tgl_max,
                                        min_value=tgl_min, max_value=tgl_max)

        df_filter = df_absensi[
            (df_absensi["Tanggal"].dt.date >= tgl_awal) &
            (df_absensi["Tanggal"].dt.date <= tgl_akhir_d)
        ].copy()

        # Heatmap kehadiran: baris = guru, kolom = tanggal
        st.markdown("#### 🗓️ Peta Kehadiran (Heatmap)")

        pivot = df_filter.pivot_table(
            index="Nama", columns="Tanggal", values="Status", aggfunc="first"
        )
        # Kode numerik untuk warna
        STATUS_NUM = {
            "Mengajar": 3, "Dinas Keluar Sekolah": 2,
            "Izin": 1, "Sakit": 1,
            "Tanpa Keterangan": 0, "Libur": -1, "Libur Tanggal Merah": -1,
        }
        pivot_num = pivot.applymap(lambda s: STATUS_NUM.get(s, -1) if pd.notna(s) else -1)

        fig_heat = go.Figure(go.Heatmap(
            z=pivot_num.values,
            x=[c.strftime("%d/%m") for c in pivot_num.columns],
            y=[n.split(",")[0] for n in pivot_num.index],
            colorscale=[
                [0.0, "#37474f"], [0.25, "#b71c1c"],
                [0.5, "#f57f17"], [0.75, "#1565c0"],
                [1.0, "#2e7d32"],
            ],
            showscale=False,
            hovertemplate="<b>%{y}</b><br>%{x}<br>Status: %{text}<extra></extra>",
            text=pivot.values,
        ))
        fig_heat.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font_color="#e0e0e0", margin=dict(t=10, b=10),
            xaxis=dict(tickangle=-45, tickfont=dict(size=9)),
            yaxis=dict(tickfont=dict(size=9)),
            height=max(400, len(pivot_num) * 22),
        )
        st.plotly_chart(fig_heat, use_container_width=True)
        st.caption("🟢 Mengajar &nbsp; 🔵 Dinas &nbsp; 🟡 Izin/Sakit &nbsp; ⚫ Tanpa Ket. &nbsp; ⬛ Libur")

        # Tabel absensi per hari
        st.markdown("#### 📋 Tabel Rekap Harian")
        rekap_harian = (
            df_filter.groupby(["Tanggal", "Status"])
            .size().reset_index(name="Jumlah Guru")
        )
        rekap_harian["Tanggal"] = rekap_harian["Tanggal"].dt.strftime("%d %b %Y")
        st.dataframe(rekap_harian, use_container_width=True, hide_index=True)

        # Download
        buf2 = BytesIO()
        df_filter[["Nama", "Tanggal", "Hari", "Status"]].to_excel(buf2, index=False)
        st.download_button(
            "⬇️ Unduh Data Absensi Harian",
            buf2.getvalue(),
            file_name="absensi_harian.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
